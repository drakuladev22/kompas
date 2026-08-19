"""Aylıq hesabat mühərriki və LOCK MEXANİZMİ — Faza 6.

Bu faylın MƏRKƏZİ testi `test_fine_inside_the_appeal_window_is_excluded`-dır:
spesifikasiya həmin qaydanı "hüquqi risk" kimi işarələyir, çünki uğurlu
etirazdan sonra işçinin premiyasından səhvən pul kəsilə bilər.
"""

from __future__ import annotations

import uuid
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.application.use_cases.reporting import (
    AttendanceRow,
    BonusPenaltyRow,
    EmployeeAttendanceFacts,
    EmployeeSalesFacts,
    MonthlyReportUseCase,
    ReportPeriod,
    ReportPeriodError,
    ReportPermissionError,
)
from src.domain.entities.fine import Fine, FineSource
from src.domain.entities.position import Position
from src.domain.value_objects.authorization import (
    PermissionEffect,
    RolePriority,
)
from src.domain.value_objects.credentials import Username
from src.domain.value_objects.identifiers import (
    EmployeeId,
    FineId,
    FineTypeId,
    PositionId,
    StoreId,
    TenantId,
)
from src.domain.value_objects.money import Money
from src.infrastructure.reporting.excel import (
    ATTENDANCE_HEADERS,
    BONUS_PENALTY_HEADERS,
    ExcelReportWriter,
)

NOW = datetime(2026, 9, 1, 10, 0, tzinfo=UTC)
TENANT = TenantId(uuid.uuid4())
EMPLOYEE = EmployeeId(uuid.uuid4())
STORE = StoreId(uuid.uuid4())
REVIEWER = EmployeeId(uuid.uuid4())


# --------------------------------------------------------------------------- #
# Aktor
# --------------------------------------------------------------------------- #


def _actor(*, can_export: bool = True):  # type: ignore[no-untyped-def]
    from src.domain.entities.employee import Employee, PermissionOverride

    position = Position(
        position_id=PositionId(uuid.uuid4()),
        code="HR_ADMIN",
        name_az="HR Admin",
        priority=RolePriority.ADMIN,
        tenant_id=TENANT,
        is_system=True,
    )
    employee = Employee(
        employee_id=EmployeeId(uuid.uuid4()),
        tenant_id=TENANT,
        position=position,
        first_name="Aygün",
        last_name="Əliyeva",
        username=Username("a.aliyeva"),
        has_password=True,
    )
    if can_export:
        employee.apply_override(
            PermissionOverride(
                flag_code="can_export_reports",
                effect=PermissionEffect.GRANT,
                granted_by=REVIEWER,
            )
        )
    return employee


def _published_fine(*, published_at: datetime, amount: str = "25.00") -> Fine:
    fine = Fine(
        fine_id=FineId(uuid.uuid4()),
        tenant_id=TENANT,
        employee_id=EMPLOYEE,
        store_id=STORE,
        source=FineSource.MANUAL_CAMERA,
        amount=Money(Decimal(amount)),
        issued_at=published_at - timedelta(hours=1),
        fine_type_id=FineTypeId(uuid.uuid4()),
        issued_by=REVIEWER,
        photo_evidence_url="https://drive/evidence.jpg",
    )
    fine.publish(reviewed_by=REVIEWER, published_at=published_at)
    return fine


def _sales_facts() -> list[EmployeeSalesFacts]:
    return [
        EmployeeSalesFacts(
            employee_id=EMPLOYEE,
            full_name="Rəşad Məmmədov",
            store_name="Bellona 28 May",
            gross_sales=Money(Decimal("18450.00")),
            earned_points=320,
        )
    ]


# --------------------------------------------------------------------------- #
# Dövr
# --------------------------------------------------------------------------- #


def test_period_key_is_zero_padded() -> None:
    assert ReportPeriod(2026, 8).key == "2026-08"


def test_period_label_is_azerbaijani() -> None:
    assert ReportPeriod(2026, 8).label_az() == "Avqust 2026"


@pytest.mark.parametrize("month", [0, 13, -1])
def test_invalid_month_is_rejected(month: int) -> None:
    with pytest.raises(ReportPeriodError):
        ReportPeriod(2026, month)


# --------------------------------------------------------------------------- #
# Səlahiyyət
# --------------------------------------------------------------------------- #


def test_export_requires_the_flag() -> None:
    use_case = MonthlyReportUseCase()
    with pytest.raises(ReportPermissionError):
        use_case.build_attendance_rows(actor=_actor(can_export=False), facts=[], now=NOW)


def test_bonus_report_also_requires_the_flag() -> None:
    use_case = MonthlyReportUseCase()
    with pytest.raises(ReportPermissionError):
        use_case.build_bonus_penalty(actor=_actor(can_export=False), facts=[], fines=[], now=NOW)


# --------------------------------------------------------------------------- #
# LOCK MEXANİZMİ (bölmə 6 — hüquqi risk)
# --------------------------------------------------------------------------- #


def test_fine_inside_the_appeal_window_is_excluded() -> None:
    """72 saat bitməyibsə cərimə bu ayın faylına DÜŞMÜR."""
    use_case = MonthlyReportUseCase()
    fine = _published_fine(published_at=NOW - timedelta(hours=10))

    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=[fine], now=NOW
    )

    row = selection.rows[0]
    assert row.confirmed_fine_count == 0
    assert row.total_fine_amount == Money(Decimal("0.00"))
    assert row.open_appeal_count == 1
    assert row.has_deferred_fines
    assert selection.included_fines == []


def test_fine_after_the_window_is_included() -> None:
    use_case = MonthlyReportUseCase()
    fine = _published_fine(published_at=NOW - timedelta(hours=100))

    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=[fine], now=NOW
    )

    row = selection.rows[0]
    assert row.confirmed_fine_count == 1
    assert row.total_fine_amount == Money(Decimal("25.00"))
    assert row.open_appeal_count == 0
    assert selection.included_fines == [fine]


def test_reversed_fine_never_enters_the_report() -> None:
    """Uğurlu etirazdan sonra premiyadan pul kəsilə BİLMƏZ."""
    use_case = MonthlyReportUseCase()
    fine = _published_fine(published_at=NOW - timedelta(hours=100))
    fine.reverse(
        decided_by=REVIEWER,
        decided_at=NOW - timedelta(hours=90),
        reason="Etiraz qəbul olundu, kamera qeydi işçini təsdiqlədi",
    )

    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=[fine], now=NOW
    )

    row = selection.rows[0]
    assert row.confirmed_fine_count == 0
    # Ləğv olunmuş cərimə "gözləyir" DEYİL — o, heç vaxt qayıtmayacaq.
    assert row.open_appeal_count == 0
    assert not row.has_deferred_fines


def test_already_exported_fine_is_not_charged_twice() -> None:
    use_case = MonthlyReportUseCase()
    fine = _published_fine(published_at=NOW - timedelta(hours=100))
    fine.mark_exported(period="2026-08", now=NOW)

    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=[fine], now=NOW
    )
    assert selection.rows[0].confirmed_fine_count == 0
    assert selection.rows[0].open_appeal_count == 0


def test_amounts_are_summed_per_employee() -> None:
    use_case = MonthlyReportUseCase()
    fines = [
        _published_fine(published_at=NOW - timedelta(hours=100), amount="25.00"),
        _published_fine(published_at=NOW - timedelta(hours=120), amount="15.50"),
    ]

    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=fines, now=NOW
    )

    assert selection.rows[0].confirmed_fine_count == 2
    assert selection.rows[0].total_fine_amount == Money(Decimal("40.50"))


def test_marking_exported_happens_only_after_the_file_is_written() -> None:
    """Ayrı metod olması qəsdəndir — yazma uğursuz olsa cərimələr açıq qalmalıdır."""
    use_case = MonthlyReportUseCase()
    fine = _published_fine(published_at=NOW - timedelta(hours=100))
    selection = use_case.build_bonus_penalty(
        actor=_actor(), facts=_sales_facts(), fines=[fine], now=NOW
    )

    assert fine.exported_period is None
    marked = use_case.mark_exported(selection=selection, period=ReportPeriod(2026, 8), now=NOW)
    assert marked == 1
    assert fine.exported_period == "2026-08"


# --------------------------------------------------------------------------- #
# Davamiyyət hesabatı — CƏRİMƏ MƏLUMATI YOXDUR
# --------------------------------------------------------------------------- #


def test_attendance_rows_carry_no_fine_columns() -> None:
    """Bölmə 6: FAYL 1-in cərimə ilə ƏLAQƏSİ YOXDUR."""
    use_case = MonthlyReportUseCase()
    rows = use_case.build_attendance_rows(
        actor=_actor(),
        facts=[
            EmployeeAttendanceFacts(
                employee_id=EMPLOYEE,
                full_name="Rəşad Məmmədov",
                store_name="Bellona 28 May",
                position_name="Satıcı",
                norm_work_days=22,
                actual_worked_days=20,
                off_days=8,
                unauthorized_absences=2,
            )
        ],
        now=NOW,
    )

    assert len(rows) == 1
    fields = set(rows[0].__dataclass_fields__)
    assert not {field for field in fields if "fine" in field or "penalty" in field}


def test_negative_counters_are_rejected() -> None:
    with pytest.raises(ReportPeriodError):
        AttendanceRow(
            employee_id=EMPLOYEE,
            full_name="X",
            store_name="Y",
            position_name="Z",
            norm_work_days=-1,
            actual_worked_days=0,
            off_days=0,
            unauthorized_absences=0,
        )


# --------------------------------------------------------------------------- #
# Excel yazıcısı
# --------------------------------------------------------------------------- #


def test_attendance_file_has_the_specified_columns(tmp_path: Path) -> None:
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_attendance(
        [
            AttendanceRow(
                employee_id=EMPLOYEE,
                full_name="Rəşad Məmmədov",
                store_name="Bellona 28 May",
                position_name="Satıcı",
                norm_work_days=22,
                actual_worked_days=20,
                off_days=8,
                unauthorized_absences=2,
            )
        ],
        period=ReportPeriod(2026, 8),
    )

    assert path.name == "Davamiyyet_2026-08.xlsx"
    sheet = load_workbook(path).active
    assert sheet is not None
    assert [cell.value for cell in sheet[1]] == list(ATTENDANCE_HEADERS)
    assert sheet["E2"].value == 22


def test_bonus_file_is_a_separate_workbook(tmp_path: Path) -> None:
    """İki hesabat QARIŞDIRILMAMALIDIR — ayrı fayl, ayrı ad."""
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_bonus_penalty(
        [
            BonusPenaltyRow(
                employee_id=EMPLOYEE,
                full_name="Rəşad Məmmədov",
                store_name="Bellona 28 May",
                gross_sales=Money(Decimal("18450.00")),
                earned_points=320,
                confirmed_fine_count=2,
                total_fine_amount=Money(Decimal("40.50")),
                open_appeal_count=1,
            )
        ],
        period=ReportPeriod(2026, 8),
    )

    assert path.name == "Premiya_Cerime_2026-08.xlsx"
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Premiya və Cərimə"]

    sheet = workbook.active
    assert sheet is not None
    assert [cell.value for cell in sheet[1]] == list(BONUS_PENALTY_HEADERS)


def test_money_cells_are_numbers_so_accountants_can_sum_them(tmp_path: Path) -> None:
    """Mətn olsaydı `SUM()` sıfır qaytarardı və səhv sükutla keçərdi."""
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_bonus_penalty(
        [
            BonusPenaltyRow(
                employee_id=EMPLOYEE,
                full_name="Rəşad Məmmədov",
                store_name="Bellona 28 May",
                gross_sales=Money(Decimal("18450.00")),
                earned_points=320,
                confirmed_fine_count=2,
                total_fine_amount=Money(Decimal("40.50")),
            )
        ],
        period=ReportPeriod(2026, 8),
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    assert isinstance(sheet["D2"].value, (int, float, Decimal))
    assert isinstance(sheet["G2"].value, (int, float, Decimal))
    assert "AZN" in str(sheet["G2"].number_format)


def test_open_appeal_state_is_shown_on_every_row(tmp_path: Path) -> None:
    """Bölmə 6: «etiraz pəncərəsi vəziyyəti (bağlı/açıq) aydın göstərilir»."""
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_bonus_penalty(
        [
            BonusPenaltyRow(
                employee_id=EMPLOYEE,
                full_name="Açıq pəncərə",
                store_name="Mağaza",
                gross_sales=Money(Decimal("100.00")),
                earned_points=1,
                confirmed_fine_count=0,
                total_fine_amount=Money(Decimal("0.00")),
                open_appeal_count=3,
            ),
            BonusPenaltyRow(
                employee_id=EmployeeId(uuid.uuid4()),
                full_name="Bağlı pəncərə",
                store_name="Mağaza",
                gross_sales=Money(Decimal("100.00")),
                earned_points=1,
                confirmed_fine_count=1,
                total_fine_amount=Money(Decimal("5.00")),
            ),
        ],
        period=ReportPeriod(2026, 8),
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    assert sheet["H2"].value == "3 cərimə gözləyir"
    assert sheet["H3"].value == "Bağlı"


def test_output_directory_is_created_when_missing(tmp_path: Path) -> None:
    target = tmp_path / "hesabatlar" / "2026"
    writer = ExcelReportWriter(output_dir=target)
    path = writer.write_attendance([], period=ReportPeriod(2026, 8))
    assert path.exists()


# --------------------------------------------------------------------------- #
# `write_table()` — ÜMUMİ cədvəl (audit jurnalının «Excel-ə İxrac Et»-i)
# --------------------------------------------------------------------------- #


def test_write_table_uses_the_header_order_not_dict_order(tmp_path: Path) -> None:
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_table(
        [{"actor": "Aygün Əliyeva", "action": "CƏRİMƏ_YARADILDI", "when": "2026-08-19"}],
        headers=[("when", "Tarix"), ("actor", "Kim"), ("action", "Əməliyyat")],
        sheet_title="Audit Jurnalı",
        file_name="Audit_2026-08.xlsx",
    )

    assert path.name == "Audit_2026-08.xlsx"
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Audit Jurnalı"]
    sheet = workbook.active
    assert sheet is not None
    assert [cell.value for cell in sheet[1]] == ["Tarix", "Kim", "Əməliyyat"]
    assert [cell.value for cell in sheet[2]] == ["2026-08-19", "Aygün Əliyeva", "CƏRİMƏ_YARADILDI"]


def test_write_table_leaves_missing_keys_blank_instead_of_raising(tmp_path: Path) -> None:
    """Audit sətirləri müxtəlif modullardan gəlir — hamısı EYNİ açarları daşımır."""
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_table(
        [{"actor": "Aygün Əliyeva"}],
        headers=[("actor", "Kim"), ("fine_amount", "Cərimə Məbləği")],
        sheet_title="Qarışıq",
        file_name="Qarisiq.xlsx",
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    assert sheet["A2"].value == "Aygün Əliyeva"
    # openpyxl boş sətri yazıb-oxuyanda `None` qaytarır (xana FİZİKİ boşdur) —
    # `KeyError` ATILMADIĞINI yoxlamaq məqsədimizdir, "" ilə `None` arasındakı
    # fərq deyil.
    assert sheet["B2"].value in ("", None)


def test_write_table_without_a_note_does_not_lose_the_last_data_row_from_the_filter(
    tmp_path: Path,
) -> None:
    """`note=""` olanda `_add_period_note` çağırılmır — avtofiltr YENƏ DƏ bütün
    sətirləri əhatə etməlidir (`_finalize`-in `footer_rows` düzəlişi)."""
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_table(
        [{"actor": "Birinci"}, {"actor": "İkinci"}],
        headers=[("actor", "Kim")],
        sheet_title="Filtr",
        file_name="Filtr.xlsx",
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    assert sheet.auto_filter.ref == "A1:A3"
    assert sheet.max_row == 3


def test_write_table_with_a_note_appends_it_below_the_data(tmp_path: Path) -> None:
    writer = ExcelReportWriter(output_dir=tmp_path)
    path = writer.write_table(
        [{"actor": "Birinci"}],
        headers=[("actor", "Kim")],
        sheet_title="Qeydli",
        file_name="Qeydli.xlsx",
        note="Bu, izah sətridir.",
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    assert sheet.cell(row=sheet.max_row, column=1).value == "Bu, izah sətridir."
    assert sheet.auto_filter.ref == "A1:A2"
