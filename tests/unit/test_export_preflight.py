"""Export təcrübəsi — kompas1.md Faza 8 (A, D, E, F, G).

Testlər ÜÇ təbəqəni ayrıca yoxlayır:

    * SAF DOMEN qaydaları (`domain/export_validation.py`) — dörd doğrulama
      qaydası, heç bir port və sahtə olmadan;
    * USE CASE (`application/use_cases/export_preflight.py`) — səlahiyyət,
      ROOT həddləri, düzəlişin tətbiqi, dövr müqayisəsi, rol filtri;
    * EXCEL yazıcısının «Qeyd» sütunu (bənd E) — real `.xlsx` yaradılır və
      geri oxunur, çünki sütunun HANSI hərfdə olduğu mühasibin düsturları
      üçün əhəmiyyətlidir.

Sahtələr BU FAYLDA yerlidir (`tests/fixtures/fakes.py`-dan yalnız `FakeClock`,
`FakeSystemLimits`, `RecordingAudit` götürülür) — eyni qərar
`test_bulk_operations_screen.py` başlığında izah olunub.
"""

from __future__ import annotations

import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Final

import pytest
from openpyxl import load_workbook

from src.application.use_cases.export_preflight import (
    EXPORT_CORRECTIONS_FLAG,
    EmployeeRosterStatus,
    ExportCorrectionPermissionError,
    ExportCorrectionReasonError,
    ExportPreflightUseCase,
    apply_corrections,
    compare_periods,
    correctable_field_options,
    filter_by_role,
    notes_for,
)
from src.application.use_cases.reporting import (
    EXPORT_REPORTS_FLAG,
    AttendanceRow,
    ReportPermissionError,
    ReportRange,
)
from src.domain.entities.employee import Employee
from src.domain.entities.position import Position
from src.domain.export_validation import (
    ExportRowFacts,
    ExportValidationCode,
    ExportValidationThresholds,
    validate_export_rows,
)
from src.domain.policies import SystemLimitKey
from src.domain.value_objects.authorization import PermissionFlag, RolePriority
from src.domain.value_objects.credentials import Username
from src.domain.value_objects.export_corrections import (
    EXPORT_TYPE_ATTENDANCE,
    ExportCorrection,
    ExportCorrectionError,
)
from src.domain.value_objects.identifiers import (
    EmployeeId,
    PositionId,
    TenantId,
    new_employee_id,
)
from src.infrastructure.reporting.excel import (
    ATTENDANCE_HEADERS,
    NOTE_HEADER,
    ExcelReportWriter,
)
from tests.fixtures.fakes import FakeClock, FakeSystemLimits, RecordingAudit

pytestmark = pytest.mark.unit

TENANT: Final = TenantId(uuid.uuid4())
NOW: Final = datetime(2026, 8, 13, 9, 0, tzinfo=UTC)
AUGUST: Final = ReportRange(date(2026, 8, 1), date(2026, 8, 31))


# --------------------------------------------------------------------------- #
# Aktorlar
# --------------------------------------------------------------------------- #


def _employee(*, flags: tuple[str, ...]) -> Employee:
    position = Position(
        position_id=PositionId(uuid.uuid4()),
        code="HR_ADMIN",
        name_az="HR Admin",
        priority=RolePriority.OPERATIONAL,
        tenant_id=TENANT,
        is_system=True,
    )
    for flag in flags:
        position.grant(PermissionFlag(code=flag, category="test"))
    return Employee(
        employee_id=EmployeeId(uuid.uuid4()),
        tenant_id=TENANT,
        position=position,
        first_name="Aysel",
        last_name="Quliyeva",
        username=Username("a.quliyeva"),
        has_password=True,
    )


def _hr_admin() -> Employee:
    """Həm hesabat çıxara, həm düzəliş edə bilən HR."""
    return _employee(flags=(EXPORT_REPORTS_FLAG, EXPORT_CORRECTIONS_FLAG))


def _reader_only() -> Employee:
    """Hesabatı GÖRÜR, lakin DÜZƏLİŞ edə BİLMİR — iki flag qəsdən ayrıdır."""
    return _employee(flags=(EXPORT_REPORTS_FLAG,))


# --------------------------------------------------------------------------- #
# Sahtələr
# --------------------------------------------------------------------------- #


class _Corrections:
    """`ExportCorrectionRepository` müqaviləsinin yaddaş nüsxəsi."""

    def __init__(self, items: list[ExportCorrection] | None = None) -> None:
        self.items = list(items or [])
        self.save_failure: Exception | None = None

    def save(self, entry: ExportCorrection) -> None:
        if self.save_failure is not None:
            raise self.save_failure
        self.items.append(entry)

    def list_for_range(
        self, tenant_id: Any, *, export_type: str, start: date, end: date
    ) -> list[ExportCorrection]:
        return [
            item
            for item in sorted(self.items, key=lambda i: (i.corrected_at, str(i.correction_id)))
            if item.export_type == export_type and start <= item.target_date <= end
        ]


class _Roster:
    """`ExportRosterProvider` sahtəsi."""

    def __init__(self, entries: list[EmployeeRosterStatus] | None = None) -> None:
        self.entries = list(entries or [])
        self.calls: list[tuple[date, date]] = []

    def roster_status(
        self, tenant_id: Any, *, start: date, end: date, store_id: Any = None
    ) -> list[EmployeeRosterStatus]:
        self.calls.append((start, end))
        return list(self.entries)


def _use_case(
    *,
    corrections: _Corrections | None = None,
    roster: _Roster | None = None,
    limits: FakeSystemLimits | None = None,
    audit: RecordingAudit | None = None,
) -> tuple[ExportPreflightUseCase, _Corrections, _Roster, RecordingAudit]:
    store = corrections or _Corrections()
    people = roster or _Roster()
    trail = audit or RecordingAudit()
    use_case = ExportPreflightUseCase(
        corrections=store,  # type: ignore[arg-type]
        roster=people,  # type: ignore[arg-type]
        audit=trail,  # type: ignore[arg-type]
        clock=FakeClock(NOW),  # type: ignore[arg-type]
        limits=limits or FakeSystemLimits(),  # type: ignore[arg-type]
    )
    return use_case, store, people, trail


def _row(
    *,
    employee_id: EmployeeId | None = None,
    name: str = "Rəşad Məmmədov",
    store: str = "Bellona 28 May",
    position: str = "Satıcı",
    norm: int = 22,
    actual: int = 20,
    off_days: int = 8,
    absences: int = 2,
) -> AttendanceRow:
    return AttendanceRow(
        employee_id=employee_id or new_employee_id(),
        full_name=name,
        store_name=store,
        position_name=position,
        norm_work_days=norm,
        actual_worked_days=actual,
        off_days=off_days,
        unauthorized_absences=absences,
    )


def _facts(
    *,
    name: str = "Rəşad Məmmədov",
    store: str = "Bellona 28 May",
    norm: int = 22,
    actual: int = 20,
    absences: int = 2,
    is_active: bool = True,
    employee_id: EmployeeId | None = None,
) -> ExportRowFacts:
    return ExportRowFacts(
        employee_id=employee_id or new_employee_id(),
        full_name=name,
        store_name=store,
        norm_work_days=norm,
        actual_worked_days=actual,
        unauthorized_absences=absences,
        is_active=is_active,
    )


_DEFAULT_THRESHOLDS: Final = ExportValidationThresholds(
    store_absence_pct=Decimal("15.0"), store_min_employees=3
)


# --------------------------------------------------------------------------- #
# (A) Dörd doğrulama qaydası — SAF DOMEN
# --------------------------------------------------------------------------- #


def test_more_worked_days_than_the_range_is_flagged() -> None:
    """Qayda 1 — 34 iş günü 31 günlük aralıqda riyazi olaraq mümkün deyil."""
    findings = validate_export_rows(
        [_facts(actual=34)], range_day_count=31, thresholds=_DEFAULT_THRESHOLDS
    )

    codes = [finding.code for finding in findings]
    assert ExportValidationCode.EXCESS_WORK_DAYS in codes
    assert "34" in findings[0].detail_az


def test_exactly_the_range_length_is_not_flagged() -> None:
    """SƏRHƏD HALI: hər gün işləmək qanunidir — bərabərlik xəbərdarlıq DEYİL."""
    findings = validate_export_rows(
        [_facts(actual=31, norm=31)], range_day_count=31, thresholds=_DEFAULT_THRESHOLDS
    )

    assert ExportValidationCode.EXCESS_WORK_DAYS not in [f.code for f in findings]


def test_inactive_employee_with_activity_is_flagged() -> None:
    """Qayda 2 — deaktiv işçi tabeldə hərəkət göstərir."""
    findings = validate_export_rows(
        [_facts(is_active=False, actual=6, absences=1)],
        range_day_count=31,
        thresholds=_DEFAULT_THRESHOLDS,
    )

    assert ExportValidationCode.INACTIVE_IN_SHEET in [f.code for f in findings]


def test_inactive_employee_without_activity_is_not_flagged() -> None:
    """Aralığın əvvəlində çıxmış işçi SƏHV DEYİL — hərəkət yoxdursa sükut."""
    findings = validate_export_rows(
        [_facts(is_active=False, actual=0, absences=0, norm=0)],
        range_day_count=31,
        thresholds=_DEFAULT_THRESHOLDS,
    )

    assert findings == ()


def test_zero_days_and_zero_absence_conflict_is_flagged() -> None:
    """Qayda 3 — plan var, hərəkət yoxdur: məlumat boşluğu."""
    findings = validate_export_rows(
        [_facts(norm=22, actual=0, absences=0)],
        range_day_count=31,
        thresholds=_DEFAULT_THRESHOLDS,
    )

    assert ExportValidationCode.ZERO_ACTIVITY_CONFLICT in [f.code for f in findings]


def test_zero_days_without_a_plan_is_not_a_conflict() -> None:
    """Planı olmayan işçi üçün «0 gün» ziddiyyət DEYİL — sadəcə plan yoxdur."""
    findings = validate_export_rows(
        [_facts(norm=0, actual=0, absences=0)],
        range_day_count=31,
        thresholds=_DEFAULT_THRESHOLDS,
    )

    assert findings == ()


def test_store_absence_anomaly_uses_the_root_threshold() -> None:
    """Qayda 4 — nisbət ROOT faizini AŞARSA mağaza işarələnir."""
    rows = [_facts(store="İstikbal Gənclik", norm=20, actual=10, absences=5) for _ in range(3)]

    findings = validate_export_rows(rows, range_day_count=31, thresholds=_DEFAULT_THRESHOLDS)

    store_findings = [f for f in findings if f.code is ExportValidationCode.STORE_ABSENCE_ANOMALY]
    assert len(store_findings) == 1
    assert store_findings[0].subject_az == "İstikbal Gənclik"
    assert store_findings[0].employee_id is None  # mağaza-səviyyəli tapıntı


def test_store_below_the_minimum_headcount_is_never_flagged() -> None:
    """Bir nəfərlik filialda 100% nisbət YALANÇI siqnaldır — hesablanmır."""
    rows = [_facts(store="Tək nəfərlik", norm=10, actual=0, absences=10)]

    findings = validate_export_rows(rows, range_day_count=31, thresholds=_DEFAULT_THRESHOLDS)

    assert ExportValidationCode.STORE_ABSENCE_ANOMALY not in [f.code for f in findings]


def test_store_without_a_plan_does_not_divide_by_zero() -> None:
    """Norma günü 0 olan mağaza sıfıra bölmə ilə ÇÖKMÜR."""
    rows = [_facts(store="Yeni filial", norm=0, actual=0, absences=3) for _ in range(4)]

    findings = validate_export_rows(rows, range_day_count=31, thresholds=_DEFAULT_THRESHOLDS)

    assert ExportValidationCode.STORE_ABSENCE_ANOMALY not in [f.code for f in findings]


def test_meaningless_thresholds_are_rejected_at_construction() -> None:
    """0 hədd HƏR mağazanı işarələyərdi — dəyər obyekti onu qəbul etmir."""
    with pytest.raises(ValueError, match="müsbət"):
        ExportValidationThresholds(store_absence_pct=Decimal("0"), store_min_employees=3)


def test_empty_row_set_produces_no_findings() -> None:
    """Rol filtri boş nəticə verəndə doğrulama ÇÖKMÜR."""
    assert validate_export_rows([], range_day_count=31, thresholds=_DEFAULT_THRESHOLDS) == ()


# --------------------------------------------------------------------------- #
# (A) Doğrulama EXPORT-U BLOKLAMIR
# --------------------------------------------------------------------------- #


def test_findings_do_not_remove_rows_from_the_export() -> None:
    """Şübhəli sətir fayla DAXİL OLUR — xəbərdarlıq bloklama deyil."""
    suspicious = _row(actual=99, norm=22)
    healthy = _row(actual=20)
    use_case, _, roster, _ = _use_case()
    roster.entries = [
        EmployeeRosterStatus(suspicious.employee_id, True, "SATICI", "Satıcı"),
        EmployeeRosterStatus(healthy.employee_id, True, "SATICI", "Satıcı"),
    ]

    review = use_case.review(
        tenant_id=TENANT,
        actor=_hr_admin(),
        rows=[suspicious, healthy],
        report_range=AUGUST,
    )

    assert review.has_findings is True
    assert len(review.rows) == 2  # heç bir sətir çıxarılmadı
    assert review.blocking_message_az() is None


def test_review_requires_the_export_flag() -> None:
    use_case, _, _, _ = _use_case()

    with pytest.raises(ReportPermissionError):
        use_case.review(
            tenant_id=TENANT,
            actor=_employee(flags=()),
            rows=[_row()],
            report_range=AUGUST,
        )


# --------------------------------------------------------------------------- #
# (D) Manual düzəliş — SƏBƏB MƏCBURİDİR
# --------------------------------------------------------------------------- #


def test_correction_without_a_reason_is_rejected() -> None:
    """Boş səbəb — nə DB-yə düşür, nə audit yazılır."""
    use_case, store, _, audit = _use_case()

    with pytest.raises(ExportCorrectionReasonError):
        use_case.record_correction(
            tenant_id=TENANT,
            actor=_hr_admin(),
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field_code="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="   ",
        )

    assert store.items == []
    assert audit.entries == []


def test_correction_with_a_too_short_reason_is_rejected_with_numbers() -> None:
    """Rədd SÜKUTLA olmur — mesaj həm faktiki, həm tələb olunan uzunluğu deyir."""
    use_case, store, _, _ = _use_case()

    with pytest.raises(ExportCorrectionReasonError) as error:
        use_case.record_correction(
            tenant_id=TENANT,
            actor=_hr_admin(),
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field_code="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="səhv",
        )

    assert "10" in error.value.user_message
    assert store.items == []


def test_root_can_raise_the_reason_minimum_above_the_database_floor() -> None:
    """ROOT dəyəri (40) DB döşəməsindən (10) yuxarıdırsa TƏTBİQ OLUNUR."""
    limits = FakeSystemLimits({SystemLimitKey.EXPORT_CORRECTION_REASON_MIN_LENGTH.value: "40"})
    use_case, store, _, _ = _use_case(limits=limits)

    with pytest.raises(ExportCorrectionReasonError):
        use_case.record_correction(
            tenant_id=TENANT,
            actor=_hr_admin(),
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field_code="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="Kassa jurnalı ilə uyğunsuzluq",  # 30 simvol — 40-dan az
        )

    assert store.items == []


def test_root_cannot_push_the_reason_minimum_below_the_database_floor() -> None:
    """ROOT 3 yazsa belə minimum 10 qalır — əks halda `INSERT` DB CHECK-inə dəyərdi."""
    limits = FakeSystemLimits({SystemLimitKey.EXPORT_CORRECTION_REASON_MIN_LENGTH.value: "3"})
    use_case, store, _, _ = _use_case(limits=limits)

    with pytest.raises(ExportCorrectionReasonError):
        use_case.record_correction(
            tenant_id=TENANT,
            actor=_hr_admin(),
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field_code="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="səhv",
        )

    assert store.items == []


def test_correction_requires_its_own_flag_not_the_export_flag() -> None:
    """Hesabatı GÖRƏN, lakin düzəliş flag-i OLMAYAN istifadəçi rədd edilir."""
    use_case, store, _, audit = _use_case()

    with pytest.raises(ExportCorrectionPermissionError):
        use_case.record_correction(
            tenant_id=TENANT,
            actor=_reader_only(),
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field_code="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="Kassa jurnalı ilə uyğunsuzluq aşkarlandı",
        )

    assert store.items == []
    assert audit.entries == []


def test_recorded_correction_is_audited_with_its_reason() -> None:
    """Audit sətri həm köhnə, həm yeni dəyəri, həm də SƏBƏBİ daşıyır."""
    use_case, store, _, audit = _use_case()
    employee_id = new_employee_id()

    entry = use_case.record_correction(
        tenant_id=TENANT,
        actor=_hr_admin(),
        employee_id=employee_id,
        target_date=date(2026, 8, 4),
        field_code="FAKTIKI_GUN",
        old_value="34",
        new_value="31",
        reason="Kassa sistemində təkrar giriş qeydi aşkarlandı",
    )

    assert store.items == [entry]
    assert audit.actions() == ["EXPORT_CORRECTION_RECORDED"]
    record = audit.entries[0]
    assert record["reason"] == "Kassa sistemində təkrar giriş qeydi aşkarlandı"
    assert record["before_state"] == {"field": "FAKTIKI_GUN", "value": "34"}
    assert record["after_state"]["value"] == "31"
    assert record["entity_id"] == entry.correction_id


def test_the_value_object_itself_refuses_an_empty_reason() -> None:
    """İKİNCİ QAPI: ekranı yan keçən skript də səbəbsiz sətir yaza bilmir."""
    with pytest.raises(ExportCorrectionError, match="səbəbi məcburidir"):
        ExportCorrection(
            tenant_id=TENANT,
            export_type=EXPORT_TYPE_ATTENDANCE,
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="   ",
            corrected_by=new_employee_id(),
            corrected_at=NOW,
        )


def test_the_value_object_refuses_an_empty_export_type() -> None:
    """Növsüz sətir heç bir faylın düzəlişi olmazdı — sükutla saxlanmır."""
    with pytest.raises(ExportCorrectionError, match="Export növü"):
        ExportCorrection(
            tenant_id=TENANT,
            export_type="   ",
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field="FAKTIKI_GUN",
            old_value="34",
            new_value="31",
            reason="Növü olmayan düzəliş cəhdi",
            corrected_by=new_employee_id(),
            corrected_at=NOW,
        )


def test_a_delta_of_zero_is_rendered_without_a_sign() -> None:
    """«0» — nə «+0», nə «−0»: sıfır fərq nişansız oxunmalıdır."""
    deltas = compare_periods([_row(absences=14)], [_row(absences=14)], significant_delta=3)

    absence = next(d for d in deltas if d.metric_key == "unauthorized_absences")
    assert absence.delta_text_az() == "0"


def test_a_negative_delta_uses_the_typographic_minus() -> None:
    deltas = compare_periods([_row(absences=11)], [_row(absences=14)], significant_delta=3)

    absence = next(d for d in deltas if d.metric_key == "unauthorized_absences")
    assert absence.delta_text_az() == "−3"


def test_list_corrections_returns_the_range_rows_without_a_full_review() -> None:
    """Kontroller düzəlişdən sonra YALNIZ jurnalı oxuya bilməlidir."""
    employee_id = new_employee_id()
    store = _Corrections([_correction(employee_id=employee_id)])
    use_case, _, _, _ = _use_case(corrections=store)

    items = use_case.list_corrections(tenant_id=TENANT, report_range=AUGUST)

    assert [item.employee_id for item in items] == [employee_id]


def test_a_correction_that_changes_nothing_is_rejected_by_the_value_object() -> None:
    """`old == new` — DB CHECK-i ilə eyni qayda, domendə də."""
    with pytest.raises(ExportCorrectionError):
        ExportCorrection(
            tenant_id=TENANT,
            export_type=EXPORT_TYPE_ATTENDANCE,
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field="FAKTIKI_GUN",
            old_value="31",
            new_value="31",
            reason="Heç nə dəyişmir, lakin yazmaq istəyirəm",
            corrected_by=new_employee_id(),
            corrected_at=NOW,
        )


def test_an_unknown_field_is_rejected_instead_of_silently_stored() -> None:
    """Tanınmayan sütun heç bir export xanasına düşməzdi — sətir rədd edilir."""
    with pytest.raises(ExportCorrectionError, match="TABEL_GUNU"):
        ExportCorrection(
            tenant_id=TENANT,
            export_type=EXPORT_TYPE_ATTENDANCE,
            employee_id=new_employee_id(),
            target_date=date(2026, 8, 4),
            field="TABEL_GUNU",
            old_value=None,
            new_value="31",
            reason="Naməlum sütun üçün düzəliş cəhdi",
            corrected_by=new_employee_id(),
            corrected_at=NOW,
        )


# --------------------------------------------------------------------------- #
# (D) Düzəliş EXPORT RƏQƏMLƏRİNƏ tətbiq olunur
# --------------------------------------------------------------------------- #


def _correction(
    *,
    employee_id: EmployeeId,
    field: str = "FAKTIKI_GUN",
    old: str | None = "34",
    new: str | None = "31",
    when: datetime = NOW,
    reason: str = "Kassa sistemində təkrar giriş qeydi aşkarlandı",
) -> ExportCorrection:
    return ExportCorrection(
        tenant_id=TENANT,
        export_type=EXPORT_TYPE_ATTENDANCE,
        employee_id=employee_id,
        target_date=date(2026, 8, 4),
        field=field,
        old_value=old,
        new_value=new,
        reason=reason,
        corrected_by=new_employee_id(),
        corrected_at=when,
    )


def test_numeric_correction_replaces_the_exported_number() -> None:
    row = _row(actual=34)

    corrected = apply_corrections([row], [_correction(employee_id=row.employee_id)])

    assert corrected[0].actual_worked_days == 31
    # Qalan sütunlar TOXUNULMAZ qalır.
    assert corrected[0].norm_work_days == row.norm_work_days


def test_the_latest_correction_wins() -> None:
    """Eyni sahəyə iki düzəliş — SONUNCU (repo sırası ilə) tətbiq olunur."""
    row = _row(actual=34)
    first = _correction(employee_id=row.employee_id, old="34", new="31")
    second = _correction(
        employee_id=row.employee_id,
        old="31",
        new="29",
        when=datetime(2026, 8, 13, 10, 0, tzinfo=UTC),
    )

    corrected = apply_corrections([row], [first, second])

    assert corrected[0].actual_worked_days == 29


def test_a_note_correction_never_touches_the_numbers() -> None:
    row = _row(actual=20)
    note = _correction(
        employee_id=row.employee_id,
        field="QEYD",
        old=None,
        new="Xəstəlik vərəqəsi gec təqdim olunub",
    )

    corrected = apply_corrections([row], [note])

    assert corrected[0] == row


def test_an_unparsable_correction_leaves_the_row_untouched() -> None:
    """«iyirmi» rəqəmə çevrilmir — BÜTÜN export dayanmır, sətir olduğu kimi qalır."""
    row = _row(actual=20)
    broken = _correction(employee_id=row.employee_id, old="20", new="iyirmi")

    corrected = apply_corrections([row], [broken])

    assert corrected[0].actual_worked_days == 20


def test_a_negative_correction_is_not_applied() -> None:
    """Mənfi dəyər `AttendanceRow` invariantını pozardı — tətbiq edilmir."""
    row = _row(actual=20)
    negative = _correction(employee_id=row.employee_id, old="20", new="-4")

    corrected = apply_corrections([row], [negative])

    assert corrected[0].actual_worked_days == 20


def test_review_validates_the_corrected_numbers_not_the_raw_ones() -> None:
    """Düzəlişdən SONRA ziddiyyət yoxdursa, xəbərdarlıq da OLMAMALIDIR."""
    row = _row(actual=99, norm=22)
    corrections = _Corrections([_correction(employee_id=row.employee_id, old="99", new="20")])
    use_case, _, roster, _ = _use_case(corrections=corrections)
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı")]

    review = use_case.review(tenant_id=TENANT, actor=_hr_admin(), rows=[row], report_range=AUGUST)

    assert review.rows[0].actual_worked_days == 20
    assert ExportValidationCode.EXCESS_WORK_DAYS not in [f.code for f in review.findings]


# --------------------------------------------------------------------------- #
# (E) «Qeyd» sütunu
# --------------------------------------------------------------------------- #


def test_notes_merge_every_correction_for_one_employee() -> None:
    employee_id = new_employee_id()
    notes = notes_for(
        [
            _correction(employee_id=employee_id, old="34", new="31"),
            _correction(employee_id=employee_id, field="QEYD", old=None, new="HR aktı 2026/114"),
        ]
    )

    assert notes[employee_id] == "Faktiki işlənilən gün: 34 → 31 | HR aktı 2026/114"


def test_attendance_file_has_the_note_column_last(tmp_path: Path) -> None:
    """Sütun SONUNCUDUR — mövcud sütunların hərfləri sürüşmür (mühasib düsturları)."""
    row = _row()
    writer = ExcelReportWriter(output_dir=tmp_path)

    path = writer.write_attendance(
        [row], period=AUGUST, notes={row.employee_id: "Faktiki işlənilən gün: 34 → 31"}
    )

    sheet = load_workbook(path).active
    assert sheet is not None
    headers = [cell.value for cell in sheet[1]]
    assert headers == list(ATTENDANCE_HEADERS)
    assert headers[-1] == NOTE_HEADER
    assert headers[4] == "Norma İş Günləri"  # E sütunu YERİNDƏ qaldı
    assert sheet["I2"].value == "Faktiki işlənilən gün: 34 → 31"


def test_attendance_file_without_notes_still_writes_the_column(tmp_path: Path) -> None:
    """`notes=None` — sütun var, xana boşdur (mövcud çağırışlar qırılmır)."""
    writer = ExcelReportWriter(output_dir=tmp_path)

    path = writer.write_attendance([_row()], period=AUGUST)

    sheet = load_workbook(path).active
    assert sheet is not None
    assert sheet["I1"].value == NOTE_HEADER
    # openpyxl boş sətri BOŞ XANA kimi saxlayır (`None`) — mühasib üçün fərq
    # yoxdur, lakin testin gözləntisi dəqiq olmalıdır.
    assert sheet["I2"].value is None


# --------------------------------------------------------------------------- #
# (F) Dövr-üzrə müqayisə
# --------------------------------------------------------------------------- #


def test_previous_range_has_the_same_length_and_ends_one_day_earlier() -> None:
    """1–15 aprel → 17–31 mart (15 gün), «keçən ay» DEYİL."""
    current = ReportRange(date(2026, 4, 1), date(2026, 4, 15))

    previous = ExportPreflightUseCase.previous_range(current)

    assert previous.start == date(2026, 3, 17)
    assert previous.end == date(2026, 3, 31)
    assert previous.day_count == current.day_count


def test_previous_range_of_a_single_day_is_yesterday() -> None:
    previous = ExportPreflightUseCase.previous_range(
        ReportRange(date(2026, 4, 10), date(2026, 4, 10))
    )

    assert (previous.start, previous.end) == (date(2026, 4, 9), date(2026, 4, 9))


def test_significant_delta_uses_the_root_threshold() -> None:
    current = [_row(absences=17)]
    previous = [_row(absences=14)]

    deltas = compare_periods(current, previous, significant_delta=3)

    absence = next(d for d in deltas if d.metric_key == "unauthorized_absences")
    assert absence.delta == 3
    assert absence.delta_text_az() == "+3"
    assert absence.is_significant is True


def test_a_delta_below_the_threshold_is_not_significant() -> None:
    deltas = compare_periods([_row(absences=15)], [_row(absences=14)], significant_delta=3)

    absence = next(d for d in deltas if d.metric_key == "unauthorized_absences")
    assert absence.is_significant is False
    assert absence.delta_text_az() == "+1"


def test_root_can_lower_the_significance_threshold(monkeypatch: Any) -> None:
    """Hədd 1 olduqda ±1 fərq də əhəmiyyətli sayılır — ədəd hardcode DEYİL."""
    limits = FakeSystemLimits({SystemLimitKey.EXPORT_PERIOD_DELTA_SIGNIFICANT.value: "1"})
    row = _row(absences=15)
    use_case, _, roster, _ = _use_case(limits=limits)
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı")]

    review = use_case.review(
        tenant_id=TENANT,
        actor=_hr_admin(),
        rows=[row],
        report_range=AUGUST,
        previous_rows=[_row(employee_id=row.employee_id, absences=14)],
    )

    absence = next(d for d in review.deltas if d.metric_key == "unauthorized_absences")
    assert absence.is_significant is True


def test_an_empty_previous_period_does_not_crash_and_shows_no_delta() -> None:
    """Keçən dövr məlumatı YOXDURSA «—» göstərilir, yalançı «−17» YOX."""
    row = _row(absences=17)
    use_case, _, roster, _ = _use_case()
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı")]

    review = use_case.review(
        tenant_id=TENANT,
        actor=_hr_admin(),
        rows=[row],
        report_range=AUGUST,
        previous_rows=[],
    )

    absence = next(d for d in review.deltas if d.metric_key == "unauthorized_absences")
    assert absence.has_baseline is False
    assert absence.delta_text_az() == "—"
    assert review.previous_range is None


def test_comparison_of_two_empty_periods_returns_zero_rows_not_an_error() -> None:
    deltas = compare_periods([], [], significant_delta=3)

    assert [d.current for d in deltas] == [0, 0, 0, 0]


# --------------------------------------------------------------------------- #
# (G) Rol filtri
# --------------------------------------------------------------------------- #


def _roster_map(*entries: EmployeeRosterStatus) -> dict[EmployeeId, EmployeeRosterStatus]:
    return {entry.employee_id: entry for entry in entries}


def test_role_filter_keeps_only_the_selected_role() -> None:
    seller = _row(name="Satıcı A")
    manager = _row(name="Menecer B")
    roster = _roster_map(
        EmployeeRosterStatus(seller.employee_id, True, "SATICI", "Satıcı"),
        EmployeeRosterStatus(manager.employee_id, True, "MAGAZA_MENECERI", "Mağaza Meneceri"),
    )

    selected = filter_by_role([seller, manager], roster=roster, role_code="SATICI")

    assert [row.full_name for row in selected] == ["Satıcı A"]


def test_an_empty_role_code_keeps_every_row() -> None:
    rows = [_row(), _row()]

    assert filter_by_role(rows, roster={}, role_code="") == rows
    assert filter_by_role(rows, roster={}, role_code=None) == rows


def test_a_role_filter_with_no_matches_does_not_crash_the_review() -> None:
    """Boş nəticə QANUNİDİR: tapıntı yoxdur, müqayisə sıfırdır, ekran çökmür."""
    row = _row()
    use_case, _, roster, _ = _use_case()
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı")]

    review = use_case.review(
        tenant_id=TENANT,
        actor=_hr_admin(),
        rows=[row],
        report_range=AUGUST,
        previous_rows=[row],
        role_code="KAMERA_NEZARETCISI",
    )

    assert review.rows == []
    assert review.findings == ()
    assert review.total_row_count == 1
    assert all(delta.current == 0 for delta in review.deltas)


def test_role_options_come_from_the_catalog_without_duplicates() -> None:
    """Seçimlər KATALOQDAN gəlir — sabit siyahı yoxdur, təkrar da yoxdur."""
    row_a, row_b = _row(), _row()
    use_case, _, roster, _ = _use_case()
    roster.entries = [
        EmployeeRosterStatus(row_a.employee_id, True, "SATICI", "Satıcı"),
        EmployeeRosterStatus(row_b.employee_id, True, "SATICI", "Satıcı"),
    ]

    review = use_case.review(
        tenant_id=TENANT, actor=_hr_admin(), rows=[row_a, row_b], report_range=AUGUST
    )

    assert [option.code for option in review.role_options] == ["SATICI"]
    assert review.role_options[0].name_az == "Satıcı"


def test_employees_without_a_role_never_become_a_filter_option() -> None:
    row = _row()
    use_case, _, roster, _ = _use_case()
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "", "—")]

    review = use_case.review(tenant_id=TENANT, actor=_hr_admin(), rows=[row], report_range=AUGUST)

    assert review.role_options == ()
    # Sətir İSƏ siyahıda qalır — rolsuz işçi export-dan düşməməlidir.
    assert len(review.rows) == 1


def test_correctable_field_options_expose_the_domain_catalog() -> None:
    codes = [code for code, _label in correctable_field_options()]

    assert codes == ["NORMA_GUN", "FAKTIKI_GUN", "OFF_DAY", "ICAZESIZ_QAYIB", "QEYD"]


# --------------------------------------------------------------------------- #
# ROOT həddləri — port yoxdursa fallback
# --------------------------------------------------------------------------- #


def test_the_anomaly_threshold_is_read_from_root() -> None:
    """Hədd 5%-ə salınanda əvvəl «normal» sayılan mağaza işarələnir."""
    limits = FakeSystemLimits({SystemLimitKey.EXPORT_STORE_ABSENCE_ANOMALY_PCT.value: "5.0"})
    rows = [_row(store="Bellona 28 May", norm=20, actual=18, absences=2) for _ in range(3)]
    use_case, _, roster, _ = _use_case(limits=limits)
    roster.entries = [
        EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı") for row in rows
    ]

    review = use_case.review(tenant_id=TENANT, actor=_hr_admin(), rows=rows, report_range=AUGUST)

    assert ExportValidationCode.STORE_ABSENCE_ANOMALY in [f.code for f in review.findings]


def test_a_broken_root_value_falls_back_instead_of_disabling_the_rule() -> None:
    """ROOT-da «abc» yazılıbsa qayda SÖNMÜR — defolt hədd işləyir."""
    limits = FakeSystemLimits({SystemLimitKey.EXPORT_STORE_ABSENCE_ANOMALY_PCT.value: "abc"})
    rows = [_row(store="Bellona 28 May", norm=10, actual=2, absences=8) for _ in range(3)]
    use_case, _, roster, _ = _use_case(limits=limits)
    roster.entries = [
        EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı") for row in rows
    ]

    review = use_case.review(tenant_id=TENANT, actor=_hr_admin(), rows=rows, report_range=AUGUST)

    assert ExportValidationCode.STORE_ABSENCE_ANOMALY in [f.code for f in review.findings]


def test_the_use_case_works_without_a_limits_port() -> None:
    """`limits=None` (maket/test yolu) — davranış defoltlarla eynidir."""
    use_case = ExportPreflightUseCase(
        corrections=_Corrections(),  # type: ignore[arg-type]
        roster=_Roster(),  # type: ignore[arg-type]
        audit=RecordingAudit(),  # type: ignore[arg-type]
        clock=FakeClock(NOW),  # type: ignore[arg-type]
    )

    review = use_case.review(
        tenant_id=TENANT, actor=_hr_admin(), rows=[_row()], report_range=AUGUST
    )

    assert len(review.rows) == 1


def test_corrections_of_another_export_type_are_not_applied() -> None:
    """Premiya faylının düzəlişi davamiyyət sətrinə TƏTBİQ OLUNMUR."""
    row = _row(actual=34)
    other = ExportCorrection(
        tenant_id=TENANT,
        export_type="BONUS_CERIME",
        employee_id=row.employee_id,
        target_date=date(2026, 8, 4),
        field="FAKTIKI_GUN",
        old_value="34",
        new_value="31",
        reason="Premiya faylına aid düzəliş sətri",
        corrected_by=new_employee_id(),
        corrected_at=NOW,
    )
    use_case, _, roster, _ = _use_case(corrections=_Corrections([other]))
    roster.entries = [EmployeeRosterStatus(row.employee_id, True, "SATICI", "Satıcı")]

    review = use_case.review(
        tenant_id=TENANT,
        actor=_hr_admin(),
        rows=[row],
        report_range=AUGUST,
        export_type=EXPORT_TYPE_ATTENDANCE,
    )

    assert review.rows[0].actual_worked_days == 34
    assert review.corrections == ()
