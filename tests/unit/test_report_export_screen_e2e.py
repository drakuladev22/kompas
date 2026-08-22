"""`ReportExportScreen` ↔ `controllers/report_export.py` — REAL Qt e2e.

──────────────────────────────────────────────────────────────────────────────
NİYƏ BU FAYL LAZIM İDİ (QA-FULL Faza 3, ikinci dalğa)
──────────────────────────────────────────────────────────────────────────────
`tests/unit/test_export_preflight_screen.py` artıq geniş örtük verir, LAKİN
İKİ TƏBƏQƏNİ AYRI sınayır: kontroller `_ScreenStub` ilə (Qt YOX), ekranın özü
isə `preflight_requested`/`export_requested`-i BİRBAŞA çağıraraq (kontroller
YOX). İkisi HEÇ VAXT birlikdə, REAL düymə kliki ilə bağlanmır — məhz bu boşluq
`test_screen_binding_coverage.py`-nin adını çəkib ÇAĞIRMADIĞI naxışdır
(CLAUDE.md §2). Bu fayl `ReportExportController.attach()`-i REAL
`ReportExportScreen`-ə bağlayır və hər sınaqda faktiki `QPushButton.click()`
işlədir.

`InlineExecutor` işlədilir (CLAUDE.md bölmə 6): fon işi test sapında SİNXRON
icra olunur, `QT_QPA_PLATFORM=offscreen` altında heç bir sap gözləməsi
lazım olmur.

Fayl sistemi sınaqları `tmp_path` ilə HƏQİQİ diskə yazır (`ExcelReportWriter`
sahtələnmir) — yalnız `_choose_output_dir` (QFileDialog) monkeypatch olunur,
CLAUDE.md §6-dakı "AYRICA metod ki, test onu Qt-siz əvəz edə bilsin" qeydinə
uyğun.
"""

from __future__ import annotations

import uuid
from contextlib import contextmanager
from datetime import UTC, date, datetime
from typing import Any, Final

import pytest
from openpyxl import load_workbook

from src.application.use_cases.export_preflight import (
    EXPORT_CORRECTIONS_FLAG,
    EmployeeRosterStatus,
    ExportPreflightReview,
    RoleOption,
)
from src.application.use_cases.reporting import (
    EXPORT_REPORTS_FLAG,
    AttendanceRow,
    ReportPermissionError,
    ReportRange,
)
from src.domain.entities.employee import Employee
from src.domain.entities.position import Position
from src.domain.export_validation import ExportValidationCode, ExportValidationFinding
from src.domain.value_objects.authorization import PermissionFlag, RolePriority
from src.domain.value_objects.credentials import Username
from src.domain.value_objects.identifiers import PositionId, TenantId, new_employee_id
from src.presentation.background_task import InlineExecutor
from src.presentation.controllers.report_export import ReportExportController
from tests.conftest import requires_qt

pytestmark = pytest.mark.unit

TENANT: Final = TenantId(uuid.uuid4())
NOW: Final = datetime(2026, 8, 21, 9, 0, tzinfo=UTC)


@pytest.fixture
def theme(qt_app):  # type: ignore[no-untyped-def]
    from src.presentation.theme.manager import ThemeManager
    from src.presentation.theme.tokens import ThemeMode

    manager = ThemeManager(preference=ThemeMode.LIGHT)
    manager.apply(qt_app)
    return manager


# --------------------------------------------------------------------------- #
# Aktorlar
# --------------------------------------------------------------------------- #


def _employee(*, flags: tuple[str, ...]) -> Employee:
    position = Position(
        position_id=PositionId(uuid.uuid4()),
        code="HR_ADMIN",
        name_az="HR Admin",
        priority=RolePriority.OPERATIONAL,
        tenant_id=TENANT,
        is_system=True,
    )
    for flag in flags:
        position.grant(PermissionFlag(code=flag, category="test"))
    return Employee(
        employee_id=new_employee_id(),
        tenant_id=TENANT,
        position=position,
        first_name="Aysel",
        last_name="Quliyeva",
        username=Username("a.quliyeva"),
        has_password=True,
    )


def _hr_admin() -> Employee:
    return _employee(flags=(EXPORT_REPORTS_FLAG, EXPORT_CORRECTIONS_FLAG))


def _reader_only() -> Employee:
    return _employee(flags=(EXPORT_REPORTS_FLAG,))


def _no_export_flag() -> Employee:
    """Menyu bu ekrana buraxmamalıdır — birbaşa naviqasiya manipulyasiyası ehtimalı."""
    return _employee(flags=())


# --------------------------------------------------------------------------- #
# Sahtələr — `report_export.py`-in HƏQİQİ sessiya müqaviləsini ödəyir
# --------------------------------------------------------------------------- #


def _row(*, name: str = "Rəşad Məmmədov", absences: int = 2) -> AttendanceRow:
    return AttendanceRow(
        employee_id=new_employee_id(),
        full_name=name,
        store_name="Bellona 28 May",
        position_name="Satıcı",
        norm_work_days=22,
        actual_worked_days=20,
        off_days=8,
        unauthorized_absences=absences,
    )


def _review(
    *,
    rows: list[AttendanceRow] | None = None,
    findings: tuple[ExportValidationFinding, ...] = (),
) -> ExportPreflightReview:
    return ExportPreflightReview(
        rows=list(rows or []),
        findings=findings,
        deltas=(),
        role_options=(RoleOption(code="SATICI", name_az="Satıcı"),),
        corrections=(),
        notes={},
        previous_range=None,
        total_row_count=len(rows or []),
    )


class _PreflightUseCase:
    """`session.export_preflight`-ın yerli əvəzi — REAL `ExportPreflightUseCase`
    YOX, çünki sınanan naxış kontroller ↔ REAL ekran körpüsüdür, use case-in öz
    daxili məntiqi YOX (o, `test_export_preflight_use_case.py`-dadır)."""

    def __init__(
        self, *, review: ExportPreflightReview | None = None, error: Exception | None = None
    ) -> None:
        self.review_result = review or _review(rows=[_row()])
        self.error = error
        self.review_calls: list[dict[str, Any]] = []

    def review(self, **kwargs: Any) -> ExportPreflightReview:
        self.review_calls.append(kwargs)
        if self.error is not None:
            raise self.error
        return self.review_result

    def record_correction(self, **kwargs: Any) -> Any:
        return object()

    @staticmethod
    def previous_range(report_range: ReportRange) -> ReportRange:
        """REAL `ExportPreflightUseCase.previous_range`-ə HƏVALƏ edilir —
        `_build_review` ONUN nəticəsini `_attendance_rows`-a ötürür, yəni
        sahtə burada `None` qaytarsaydı hər preflight çağırışı `AttributeError`
        atardı (sınananın ÖZÜ deyil, sahtənin qüsuru)."""
        from src.application.use_cases.export_preflight import ExportPreflightUseCase

        return ExportPreflightUseCase.previous_range(report_range)


class _Reports:
    """`session.reports`-ın yerli əvəzi. `resolve_month`/`resolve_range` REAL
    `MonthlyReportUseCase`-ə HƏVALƏ edilir — `ReportRange` invariantları
    (əks aralıq, il yoxlaması) məhz orada yaşayır."""

    def __init__(self, rows: list[AttendanceRow] | None = None, *, limits: Any = None) -> None:
        self.rows = list(rows if rows is not None else [_row()])
        self.marked: list[tuple[Any, str]] = []
        from src.application.use_cases.reporting import MonthlyReportUseCase

        self._real = MonthlyReportUseCase(limits=limits)

    def build_attendance_rows_for_range(self, **kwargs: Any) -> list[AttendanceRow]:
        return list(self.rows)

    def resolve_month(self, *, year: int, month: int) -> ReportRange:
        return self._real.resolve_month(year=year, month=month)

    def resolve_range(self, *, tenant_id: Any, start: date, end: date) -> ReportRange:
        return self._real.resolve_range(tenant_id=tenant_id, start=start, end=end)


class _ReportFacts:
    def attendance_facts(self, tenant_id: Any, **kwargs: Any) -> list[Any]:
        return []

    def plan_facts(self, tenant_id: Any, **kwargs: Any) -> list[Any]:
        return []

    def sales_facts(self, tenant_id: Any, **kwargs: Any) -> list[Any]:
        return []


class _FineRepo:
    def list_in_range(self, tenant_id: Any, **kwargs: Any) -> list[Any]:
        return []


class _Uow:
    def __init__(self) -> None:
        self.fines = _FineRepo()


class _RosterRepo:
    def roster_status(self, tenant_id: Any, **kwargs: Any) -> list[EmployeeRosterStatus]:
        return [EmployeeRosterStatus(new_employee_id(), True, "SATICI", "Satıcı")]


class _Limits:
    def get_int(self, tenant_id: Any, key: str, default: int) -> int:
        return default


class _Session:
    def __init__(self, *, preflight: _PreflightUseCase, reports: _Reports | None = None) -> None:
        self.tenant_id = TENANT
        self.export_preflight = preflight
        self.reports = reports or _Reports()
        self.report_facts = _ReportFacts()
        self.export_roster = _RosterRepo()
        self.uow = _Uow()
        self.limits = _Limits()
        self.commits = 0

    def commit(self) -> None:
        self.commits += 1


class _Context:
    """`ApplicationContext.session()` müqaviləsinin yerli əvəzi — hər çağırış
    YENİ sessiya sayır (CLAUDE.md §6: "Kontroller sessiyanı SAXLAMIR")."""

    def __init__(self, session_factory: Any) -> None:
        self._factory = session_factory
        self.sessions: list[_Session] = []
        self.tenant_id = TENANT

    @contextmanager
    def session(self, *, user_id: Any = None) -> Any:
        created = self._factory()
        self.sessions.append(created)
        yield created


def _controller(context: _Context, actor: Employee) -> ReportExportController:
    return ReportExportController(context, actor, executor=InlineExecutor())  # type: ignore[arg-type]


def _screen(theme: Any, qtbot: Any) -> Any:
    from src.presentation.screens.group_h import ReportExportScreen

    screen = ReportExportScreen(theme)
    qtbot.addWidget(screen)
    return screen


def _click_card(screen: Any, index: int) -> None:
    """`_REPORT_CARDS`-ın SIRASI: 0 = davamiyyət, 1 = premiya/cərimə.

    Hər iki kart eyni «Doğrula və Hazırla» mətnini daşıyır — mətnlə YOX,
    yerləşmə sırası ilə seçilir (kartlar `cards_layout`-a `_REPORT_CARDS`
    sırası ilə əlavə olunur).
    """
    from PySide6.QtWidgets import QPushButton

    buttons = [b for b in screen.findChildren(QPushButton) if b.text() == "Doğrula və Hazırla"]
    buttons[index].click()


# --------------------------------------------------------------------------- #
# 1. Real kart kliki — doğrulama REAL ekranda görünür
# --------------------------------------------------------------------------- #


@requires_qt
def test_the_real_attendance_card_click_runs_preflight_and_renders_real_rows_and_findings(
    qtbot, theme
) -> None:  # type: ignore[no-untyped-def]
    row = _row(name="Kamran Vəliyev")
    review = _review(
        rows=[row],
        findings=(
            ExportValidationFinding(
                code=ExportValidationCode.EXCESS_WORK_DAYS,
                subject_az="Kamran Vəliyev",
                detail_az="34 iş günü qeydə alınıb",
                employee_id=row.employee_id,
            ),
        ),
    )
    preflight = _PreflightUseCase(review=review)
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)

    _controller(context, _hr_admin()).attach(screen)
    _click_card(screen, 0)  # REAL düymə — sahtə deyil

    assert screen._preflight_card.isVisibleTo(screen) is True
    assert screen._finding_table.row_count == 1
    assert screen._confirm_button.isEnabled() is True, "xəbərdarlıq təsdiqi BLOKLAMAMALIDIR"
    assert screen.active_report() == "attendance"


# --------------------------------------------------------------------------- #
# 2. Real təsdiq düyməsi — doğrulama işlədilməyibsə HEÇ NƏ olmur
# --------------------------------------------------------------------------- #


@requires_qt
def test_the_confirm_button_does_nothing_before_any_preflight_has_run(qtbot, theme) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase()
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)
    # `attach()` ÖZÜ artıq rol kataloqu üçün BİR sessiya açır (`_load_role_
    # options`) — buradan sonra YENİ sessiya AÇILMAMALIDIR.
    sessions_after_attach = len(context.sessions)

    screen._confirm_button.click()  # REAL klik — heç bir doğrulama edilməyib

    assert len(context.sessions) == sessions_after_attach
    assert preflight.review_calls == []


# --------------------------------------------------------------------------- #
# 3. Real təsdiq → HƏQİQİ fayl diskə yazılır
# --------------------------------------------------------------------------- #


@requires_qt
def test_confirming_after_preflight_writes_a_real_file_to_disk(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    monkeypatch.setattr(
        ReportExportController, "_choose_output_dir", lambda self, _screen: tmp_path
    )
    _controller(context, _hr_admin()).attach(screen)

    _click_card(screen, 0)
    screen._confirm_button.click()  # REAL «Təsdiqlə və Export Et»

    written = list(tmp_path.glob("Davamiyyet_*.xlsx"))
    assert len(written) == 1
    workbook = load_workbook(written[0])
    # `_add_period_note` HƏR fayla İKİ ƏLAVƏ sətir yazır (boş ayırıcı + izah,
    # bax `excel.py::_add_period_note`) — max_row = başlıq(1) + sətir(1) + 2.
    assert workbook.active.max_row == 4
    assert workbook.active.cell(row=2, column=2).value == "Rəşad Məmmədov"
    assert "yazıldı" in screen._preflight_message.text()
    assert any(s.commits >= 1 for s in context.sessions)


# --------------------------------------------------------------------------- #
# 4. Fayl sistemi — hədəf qovluq YERİNDƏ FAYL ilə BLOKLANIB
# --------------------------------------------------------------------------- #


@requires_qt
def test_a_target_path_blocked_by_an_existing_file_shows_a_clear_message_and_does_not_crash(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    """`_output_dir.mkdir(parents=True, exist_ok=True)` yolun sonu FAYLDIRSA
    `exist_ok` onu XİLAS ETMİR — `FileExistsError` atılır. Real ssenari:
    seçilmiş qovluq bu arada silinib, yerinə eyni adda fayl yaranıb."""
    blocked = tmp_path / "hesabatlar"
    blocked.write_bytes(b"bu qovluq DEYIL - adi fayldir")

    preflight = _PreflightUseCase(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    monkeypatch.setattr(ReportExportController, "_choose_output_dir", lambda self, _screen: blocked)
    _controller(context, _hr_admin()).attach(screen)

    _click_card(screen, 0)
    screen._confirm_button.click()  # ÇÖKMƏMƏLİDİR

    message = screen._preflight_message.text()
    assert "yazıla bilmədi" in message
    assert blocked.read_bytes() == b"bu qovluq DEYIL - adi fayldir", "orijinal fayl TOXUNULMAYIB"


# --------------------------------------------------------------------------- #
# 5. Fayl sistemi — köhnə/xarab faylın ÜSTÜNDƏN yazma
# --------------------------------------------------------------------------- #


@requires_qt
def test_exporting_over_a_stale_existing_file_overwrites_it_cleanly(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    """Faylın adı DÖVR AÇARINDAN gəlir — eyni ayın ikinci export-u eyni fayla
    düşür. Köhnə fayl HƏTTA yararsız (korlanmış) olsa belə üzərinə yazılmalı,
    "artıq mövcuddur" xətası ÇIXMAMALIDIR."""
    stale = tmp_path / "Davamiyyet_2026-08.xlsx"
    stale.write_bytes(b"bu HEC bir zaman duzgun xlsx deyil")

    preflight = _PreflightUseCase(review=_review(rows=[_row(name="Yeni Sətir")]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    monkeypatch.setattr(
        ReportExportController, "_choose_output_dir", lambda self, _screen: tmp_path
    )
    _controller(context, _hr_admin()).attach(screen)

    _click_card(screen, 0)
    screen._confirm_button.click()  # ÇÖKMƏMƏLİDİR

    workbook = load_workbook(stale)
    assert workbook.active.cell(row=2, column=2).value == "Yeni Sətir"
    assert len(list(tmp_path.glob("Davamiyyet_*.xlsx"))) == 1, "ikinci fayl YARANMAMALIDIR"


# --------------------------------------------------------------------------- #
# 6. Boş nəticə — 0 sətir belə aydın mesajla, keçərli faylla bitir
# --------------------------------------------------------------------------- #


@requires_qt
def test_a_zero_row_result_still_writes_a_valid_file_with_a_clear_zero_row_message(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(review=_review(rows=[]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    monkeypatch.setattr(
        ReportExportController, "_choose_output_dir", lambda self, _screen: tmp_path
    )
    _controller(context, _hr_admin()).attach(screen)

    _click_card(screen, 0)
    assert screen._finding_table.isVisibleTo(screen) is False
    assert "tapılmadı" in screen._preflight_message.text()

    screen._confirm_button.click()

    written = list(tmp_path.glob("Davamiyyet_*.xlsx"))
    assert len(written) == 1, "boş nəticə də KEÇƏRLİ fayl yaratmalıdır"
    workbook = load_workbook(written[0])
    # başlıq(1) + 0 sətir + 2 ayırıcı/izah sətri (`_add_period_note`) = 3.
    assert workbook.active.max_row == 3, "yalnız başlıq və izah sətri — sıfır məlumat sətri"
    assert workbook.active.cell(row=2, column=1).value is None, "boş ayırıcı sətir, məlumat YOX"
    assert "0 sətir" in screen._preflight_message.text()


# --------------------------------------------------------------------------- #
# 7. Real «Aralığı Tətbiq Et» — əks aralıq domendə RƏDD edilir
# --------------------------------------------------------------------------- #


@requires_qt
def test_the_real_apply_range_button_rejects_an_inverted_range_before_any_session_writes(
    qtbot, theme
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase()
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)

    screen._range_mode.setCurrentIndex(1)  # `[Xüsusi Aralıq]`
    screen._range_start.set_text("2026-04-15")
    screen._range_end.set_text("2026-04-01")  # bitmə < başlanğıc
    screen._on_range_applied()  # REAL «Aralığı Tətbiq Et» məntiqi

    assert screen._range_message.isVisibleTo(screen) is True
    assert "əvvəl ola bilməz" in screen._range_message.text()
    assert preflight.review_calls == []


# --------------------------------------------------------------------------- #
# 8. Real «Aralığı Tətbiq Et» — ROOT həddini aşan aralıq
# --------------------------------------------------------------------------- #


@requires_qt
def test_the_real_apply_range_button_shows_the_root_limit_message_for_a_too_wide_range(
    qtbot, theme
) -> None:  # type: ignore[no-untyped-def]
    from src.domain.policies import SystemLimitKey
    from tests.fixtures.fakes import FakeSystemLimits

    limits = FakeSystemLimits({SystemLimitKey.REPORT_RANGE_MAX_DAYS.value: "31"})
    preflight = _PreflightUseCase()
    context = _Context(lambda: _Session(preflight=preflight, reports=_Reports(limits=limits)))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)

    screen._range_mode.setCurrentIndex(1)
    screen._range_start.set_text("2026-01-01")
    screen._range_end.set_text("2026-12-31")
    screen._on_range_applied()

    message = screen._range_message.text()
    assert "31" in message and "365" in message
    assert preflight.review_calls == []
    assert screen._preflight_message.text() == "", "səs-küy doğrulama kartına YAZILMIR"


# --------------------------------------------------------------------------- #
# 9. Hər əməliyyat ÖZ sessiyasını açır və commit edir
# --------------------------------------------------------------------------- #


@requires_qt
def test_each_real_click_opens_and_commits_its_own_session(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    monkeypatch.setattr(
        ReportExportController, "_choose_output_dir", lambda self, _screen: tmp_path
    )
    _controller(context, _hr_admin()).attach(screen)

    before_attach = len(context.sessions)  # `attach()` özü artıq `_load_role_options` üçün açır
    _click_card(screen, 0)
    after_preflight = len(context.sessions)
    screen._confirm_button.click()
    after_export = len(context.sessions)

    assert after_preflight > before_attach
    assert after_export > after_preflight, "export ÖZ AYRICA sessiyasını açmalıdır"
    assert all(s.commits == 1 for s in context.sessions), "hər sessiya BİR dəfə commit edir"


# --------------------------------------------------------------------------- #
# 10 & 11. Domen/gözlənilməz xəta — SÜKUTLA UDULMUR
# --------------------------------------------------------------------------- #


@requires_qt
def test_a_domain_permission_error_is_shown_clearly_not_swallowed(qtbot, theme) -> None:  # type: ignore[no-untyped-def]
    """Menyu bu ekranı bağlamalıdır, LAKİN use case özü də fail-closed-dur —
    ikinci qat sükutla keçməməlidir."""
    preflight = _PreflightUseCase(
        error=ReportPermissionError("flag yoxdur", context={"actor_id": "x"})
    )
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _no_export_flag()).attach(screen)

    _click_card(screen, 0)  # ÇÖKMƏMƏLİDİR

    assert screen._preflight_card.isVisibleTo(screen) is True
    assert "səlahiyyətiniz yoxdur" in screen._preflight_message.text()


@requires_qt
def test_an_unexpected_exception_never_leaks_technical_detail_to_the_real_screen(
    qtbot, theme
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(error=RuntimeError("psycopg connection reset by peer"))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)

    _click_card(screen, 0)  # ÇÖKMƏMƏLİDİR

    message = screen._preflight_message.text()
    assert "psycopg" not in message
    assert message  # ekran heç nə göstərmədən susmur


# --------------------------------------------------------------------------- #
# 12. GÖRMƏK = SƏLAHİYYƏTİN OLMASI — real ekranda
# --------------------------------------------------------------------------- #


@requires_qt
def test_the_corrections_section_is_completely_hidden_for_an_actor_without_the_flag(
    qtbot, theme
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)

    _controller(context, _reader_only()).attach(screen)

    assert screen._corrections_card.isVisibleTo(screen) is False
    _click_card(screen, 0)
    assert screen._corrections_card.isVisibleTo(screen) is False, (
        "doğrulamadan sonra da GİZLİ qalmalıdır"
    )


# --------------------------------------------------------------------------- #
# 13 & 14. Real düzəliş dialoqu — ekstremal/az girişlə
# --------------------------------------------------------------------------- #


def _open_correction_dialog_and_submit(
    screen: Any,
    monkeypatch: pytest.MonkeyPatch,
    *,
    value: str,
    reason: str,
) -> None:
    from PySide6.QtWidgets import QPushButton

    from src.presentation.screens.group_h import ExportCorrectionDialog

    def fake_exec(self: ExportCorrectionDialog) -> int:
        self._value.set_text(value)
        self._reason.setPlainText(reason)
        submit = next(b for b in self.findChildren(QPushButton) if b.text() == "Düzəlişi Yaz")
        submit.click()
        return 0

    monkeypatch.setattr(ExportCorrectionDialog, "exec", fake_exec)
    screen._correction_button.click()  # REAL «Düzəliş Et»


@requires_qt
def test_hostile_and_oversized_correction_input_passes_through_the_real_dialog_without_crashing(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch
) -> None:  # type: ignore[no-untyped-def]
    recorded: list[dict[str, Any]] = []

    class _RecordingPreflight(_PreflightUseCase):
        def record_correction(self, **kwargs: Any) -> Any:
            recorded.append(kwargs)
            return object()

    preflight = _RecordingPreflight(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)
    _click_card(screen, 0)  # `_active_report` = "attendance", düzəliş bölməsi doldurulur

    hostile = "'; DROP TABLE fines; --\n" + "🔥" * 20 + "A" * 10_000
    _open_correction_dialog_and_submit(
        screen, monkeypatch, value=hostile, reason="Kassa sistemində təkrar giriş " * 3
    )

    assert len(recorded) == 1
    assert recorded[0]["new_value"] == hostile


@requires_qt
def test_an_under_length_reason_is_rejected_by_the_real_dialog_before_any_write(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch
) -> None:  # type: ignore[no-untyped-def]
    from PySide6.QtWidgets import QPushButton

    from src.presentation.screens.group_h import ExportCorrectionDialog

    recorded: list[dict[str, Any]] = []

    class _RecordingPreflight(_PreflightUseCase):
        def record_correction(self, **kwargs: Any) -> Any:
            recorded.append(kwargs)
            return object()

    preflight = _RecordingPreflight(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    _controller(context, _hr_admin()).attach(screen)
    _click_card(screen, 0)

    def fake_exec(self: ExportCorrectionDialog) -> int:
        self._value.set_text("18")
        self._reason.setPlainText("qısa")  # ROOT minimumundan az
        submit = next(b for b in self.findChildren(QPushButton) if b.text() == "Düzəlişi Yaz")
        submit.click()
        assert self._reason_error.isVisibleTo(self) is True
        return self.reject()

    monkeypatch.setattr(ExportCorrectionDialog, "exec", fake_exec)
    screen._correction_button.click()  # ÇÖKMƏMƏLİDİR

    assert recorded == []


# --------------------------------------------------------------------------- #
# 15. SÜRƏTLİ İKİQAT KLİK — DÜZƏLDİLDİ (`ui` sahibi, `_start_task` qapısı)
# --------------------------------------------------------------------------- #


@requires_qt
def test_a_second_click_while_the_first_export_is_still_executing_is_not_rejected(
    qtbot, theme, monkeypatch: pytest.MonkeyPatch, tmp_path: Any
) -> None:  # type: ignore[no-untyped-def]
    preflight = _PreflightUseCase(review=_review(rows=[_row()]))
    context = _Context(lambda: _Session(preflight=preflight))
    screen = _screen(theme, qtbot)
    controller = _controller(context, _hr_admin())
    controller.attach(screen)
    _click_card(screen, 0)  # `_active_report` hazırlanır — kliklərdən ƏVVƏL, sahtələnmədən

    monkeypatch.setattr(
        ReportExportController, "_choose_output_dir", lambda self, _screen: tmp_path
    )

    calls: list[str] = []
    original_build_review = ReportExportController._build_review

    def _reentrant_build_review(self: ReportExportController, inputs: Any, report_key: str) -> Any:
        calls.append(report_key)
        if len(calls) == 1:
            # Biz İNDİ birinci tapşırığın `job()`-unun İÇİNDƏYİK — REAL
            # istehsalatda (`QtPoolExecutor`) bura məhz fon sapı çatanda
            # düşülür. `InlineExecutor` sinxron olduğu üçün `_start_task`
            # `self._task`-ı `.run(job)` çağırmazdan ƏVVƏL yazır (bax onun
            # şərhi) — ona görə bu reentrant klik artıq DOĞRU (running)
            # `self._task`-ı görür və qapıdan REDD edilir.
            screen._confirm_button.click()  # REAL ikinci klik, İCRA HƏLƏ BİTMƏYİB
        return original_build_review(self, inputs, report_key)

    monkeypatch.setattr(ReportExportController, "_build_review", _reentrant_build_review)

    screen._confirm_button.click()  # REAL birinci klik

    # DOĞRU DAVRANIŞ: ikinci (iç-içə) klik REDD edilib, `_build_review`
    # CƏMİ BİR dəfə çağırılıb (`_start_task` qapısı, `ui` sahibi).
    assert len(calls) == 1, "ikinci klik REDD edilməli idi — is_running yoxlaması yoxdur"
