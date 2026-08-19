"""Excel Export Engine — İKİ AYRI `.xlsx` faylı (spesifikasiya bölmə 6).

──────────────────────────────────────────────────────────────────────────────
İKİ FAYL, İKİ METOD — BİR "HAMISI BİR YERDƏ" METODU YOXDUR
──────────────────────────────────────────────────────────────────────────────
`write_attendance()` və `write_bonus_penalty()` AYRI metodlardır və ortaq
"hər ikisini yaz" metodu QƏSDƏN yoxdur. Bölmə 6: iki fayl "ayrı məqsədə
xidmət edir və QARIŞDIRILMAMALIDIR". Bir metod hər iki vərəqi eyni iş
kitabına yazsaydı, mühasib cərimə sütununu maaş cədvəlinin yanında görərdi —
məhz spesifikasiyanın bağladığı səhv.

──────────────────────────────────────────────────────────────────────────────
MƏBLƏĞ NİYƏ `Decimal` KİMİ YAZILIR, MƏTN KİMİ YOX
──────────────────────────────────────────────────────────────────────────────
Mühasib faylda cəm çıxarır. Məbləğ "12.50 AZN" mətni olsaydı, `SUM()`
sıfır qaytarardı və səhv sükutla keçərdi. Ona görə hüceyrə ƏDƏD-dir,
"AZN" isə format maskasındadır (`#,##0.00" AZN"`).

──────────────────────────────────────────────────────────────────────────────
FAYL ADI NİYƏ DÖVR AÇARINI DAŞIYIR
──────────────────────────────────────────────────────────────────────────────
`Davamiyyet_2026-08.xlsx` — iki fərqli ayın faylı eyni qovluqda qarışmasın.
Latın hərfləri ilə: fayl adı Windows/e-poçt/1C arasında gəzir və `ə/ı/ş`
bəzi köhnə sistemlərdə pozulur. Fayl DAXİLİ isə tam Azərbaycan dilindədir.

──────────────────────────────────────────────────────────────────────────────
«QEYD» SÜTUNU NİYƏ SONUNCUDUR VƏ NİYƏ BOŞ OLA BİLƏR (kompas1.md Faza 8, E)
──────────────────────────────────────────────────────────────────────────────
Sütun hər iki fayla ƏLAVƏ edilib, mövcud sütunların SIRASI isə DƏYİŞMƏYİB:
mühasibin `SUM(G:G)` kimi hazır düsturları, 1C idxal şablonları və pul
formatının bağlı olduğu hərflər (D, G) yerində qalmalıdır. Yeni sütunu ortaya
salmaq hamısını sükutla sürüşdürərdi.

Məzmun İSTƏYƏ BAĞLIDIR (`notes=None` → sütun var, xanalar boş). Mənbəyi
`export_manual_corrections`-dır (`export_preflight.notes_for()`) — yəni faylda
görünən hər qeydin arxasında müəllif, vaxt və MƏCBURİ səbəb var. Ayrıca
"sərbəst qeyd" sahəsi QƏSDƏN yaradılmadı: audit-siz ikinci mətn kanalı
"düzəldilib, amma niyə bilinmir" halını geri gətirərdi.
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.shared.exceptions import KompasOSError
from src.shared.logger import LogChannel, get_logger

if TYPE_CHECKING:
    from collections.abc import Mapping

    from src.application.use_cases.reporting import (
        AttendanceRow,
        BonusPenaltyRow,
        ReportPeriod,
        ReportRange,
    )
    from src.domain.value_objects.identifiers import EmployeeId

    #: Fayl adı və başlıq sətri üçün lazım olan YEGANƏ müqavilə: `key` + `label_az()`.
    #: `ReportPeriod` (tam ay) və `ReportRange` (Faza 7, xüsusi aralıq) ikisi də
    #: onu ödəyir. Yazıcı hansı formanın seçildiyini BİLMƏMƏLİDİR — bilsəydi,
    #: hər yeni dövr tipi bu modulda `if` budağı tələb edərdi.
    ExportPeriod = ReportPeriod | ReportRange

_audit_log = get_logger(__name__, channel=LogChannel.AUDIT)

#: Hər iki faylın SONUNCU sütunu (kompas1.md Faza 8, bənd E) — bax modul
#: başlığı. Sabit ayrıca elan olunur ki, iki başlıq siyahısında mətn
#: fərqlənməsin ("Qeyd" ↔ "Qeydlər" kimi sürüşmə auditdə çaşqınlıq yaradardı).
NOTE_HEADER: Final = "Qeyd"

#: FAYL 1 sütunları — bölmə 6-dakı ardıcıllıqla.
ATTENDANCE_HEADERS: Final[tuple[str, ...]] = (
    "İşçi ID",
    "Ad Soyad",
    "Mağaza",
    "Vəzifə",
    "Norma İş Günləri",
    "Faktiki İşlənilən Gün Sayı",
    "Off-Day Sayı",
    "İcazəsiz Qayıb",
    NOTE_HEADER,
)

#: FAYL 2 sütunları — bölmə 6-dakı ardıcıllıqla.
#: Sonuncu sütun spesifikasiyada sadalanmır, lakin oradakı tələbdən doğur:
#: "hər sətir üçün «etiraz pəncərəsi vəziyyəti» (bağlı/açıq) aydın göstərilir".
BONUS_PENALTY_HEADERS: Final[tuple[str, ...]] = (
    "İşçi ID",
    "Ad Soyad",
    "Mağaza",
    "Brutto Satış Məbləği",
    "Qazanılan Satış Xalları",
    "Təsdiqlənmiş Cərimə Sayı",
    "Premiyadan Tutulacaq Yekun Cərimə Məbləği",
    "Etiraz Pəncərəsi",
    NOTE_HEADER,
)

_MONEY_FORMAT: Final[str] = '#,##0.00" AZN"'
_HEADER_FILL = PatternFill("solid", fgColor="0B1D3A")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MIN_COLUMN_WIDTH: Final[int] = 12
_MAX_COLUMN_WIDTH: Final[int] = 46


class ExcelExportError(KompasOSError):
    """Fayl yazıla bilmədi."""

    user_message = "Hesabat faylı yazıla bilmədi. Qovluğa yazma icazəsini yoxlayın."


class ExcelReportWriter:
    """İki aylıq hesabatı `.xlsx` kimi yazır."""

    def __init__(self, *, output_dir: Path) -> None:
        self._output_dir = output_dir

    # ------------------------------- FAYL 1 ---------------------------------- #

    def write_attendance(
        self,
        rows: list[AttendanceRow],
        *,
        period: ExportPeriod,
        notes: Mapping[EmployeeId, str] | None = None,
    ) -> Path:
        """Aylıq Davamiyyət Hesabatı — YALNIZ fiks maaş üçün.

        `notes` DEFOLT `None` — mövcud çağırış yerləri (və testlər) DƏYİŞMƏDƏN
        işləməyə davam edir; sütun yaranır, xanalar boş qalır (bax modul
        başlığı).
        """
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:  # pragma: no cover — openpyxl həmişə vərəq yaradır
            raise ExcelExportError("Boş iş kitabı yaradıla bilmədi")
        sheet.title = "Davamiyyət"

        lookup = notes or {}
        self._write_headers(sheet, ATTENDANCE_HEADERS)
        for row in rows:
            sheet.append(
                [
                    str(row.employee_id),
                    row.full_name,
                    row.store_name,
                    row.position_name,
                    row.norm_work_days,
                    row.actual_worked_days,
                    row.off_days,
                    row.unauthorized_absences,
                    lookup.get(row.employee_id, ""),
                ]
            )

        self._add_period_note(
            sheet,
            column_count=len(ATTENDANCE_HEADERS),
            note=(
                f"{period.label_az()} — yalnız fiks maaş uçotu üçün. "
                f"Cərimə məlumatı bu faylda YOXDUR."
            ),
        )
        self._finalize(sheet, len(ATTENDANCE_HEADERS))
        return self._save(workbook, f"Davamiyyet_{period.key}.xlsx", report="attendance")

    # ------------------------------- FAYL 2 ---------------------------------- #

    def write_bonus_penalty(
        self,
        rows: list[BonusPenaltyRow],
        *,
        period: ExportPeriod,
        notes: Mapping[EmployeeId, str] | None = None,
    ) -> Path:
        """Premiya & Cərimə Hesabatı — cərimələr PREMİYADAN tutulur."""
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:  # pragma: no cover
            raise ExcelExportError("Boş iş kitabı yaradıla bilmədi")
        sheet.title = "Premiya və Cərimə"

        lookup = notes or {}
        self._write_headers(sheet, BONUS_PENALTY_HEADERS)
        for row in rows:
            sheet.append(
                [
                    str(row.employee_id),
                    row.full_name,
                    row.store_name,
                    row.gross_sales.amount,
                    row.earned_points,
                    row.confirmed_fine_count,
                    row.total_fine_amount.amount,
                    (
                        f"{row.open_appeal_count} cərimə gözləyir"
                        if row.has_deferred_fines
                        else "Bağlı"
                    ),
                    lookup.get(row.employee_id, ""),
                ]
            )

        # Pul sütunları: D (brutto satış) və G (yekun cərimə).
        for column in ("D", "G"):
            for cell in sheet[column][1:]:
                cell.number_format = _MONEY_FORMAT

        self._add_period_note(
            sheet,
            column_count=len(BONUS_PENALTY_HEADERS),
            note=(
                f"{period.label_az()} — cərimələr ƏSAS MAAŞDAN DEYİL, aylıq "
                f"PREMİYADAN kəsilir. Etiraz pəncərəsi açıq olan və ləğv "
                f"edilmiş cərimələr bu fayla DAXİL EDİLMƏYİB."
            ),
        )
        self._finalize(sheet, len(BONUS_PENALTY_HEADERS))
        return self._save(workbook, f"Premiya_Cerime_{period.key}.xlsx", report="bonus_penalty")

    # ------------------------------- ÜMUMİ CƏDVƏL ----------------------------- #

    def write_table(
        self,
        rows: list[dict[str, str]],
        *,
        headers: list[tuple[str, str]],
        sheet_title: str,
        file_name: str,
        note: str = "",
    ) -> Path:
        """İXTİYARİ cədvəl — audit jurnalı «Excel-ə İxrac Et» düyməsi üçün (bax
        modul başlığı: `write_attendance`/`write_bonus_penalty` YALNIZ ÖZ sabit
        sxemlərini bilir, audit sətirləri isə modul-modul fərqli sahələr daşıyır).

        `headers`: `(açar, başlıq_az)` cütləri — açar `rows`-dakı `dict`-in
        açarıdır, başlıq isə Excel-də görünən mətndir. Sıra `headers`-in
        SIRASIDIR (`dict` sırası deyil) ki, çağıran sütun ardıcıllığını idarə
        edə bilsin.

        `rows`-da açar ÇATIŞMIRSA `KeyError` ATILMIR, boş xana yazılır: audit
        sətirləri MÜXTƏLİF modullardan (cərimə, icazə, tapşırıq, ...) gəlir
        və hamısı EYNİ sütun dəstini doldurmaya bilər — məs. «Cərimə Məbləği»
        sütunu icazə hadisəsində mənasızdır, xəta yox, sadəcə boş xana.
        """
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:  # pragma: no cover — openpyxl həmişə vərəq yaradır
            raise ExcelExportError("Boş iş kitabı yaradıla bilmədi")
        # Excel vərəq adı 31 simvoldan uzun ola BİLMƏZ (openpyxl bunu ATIR) —
        # audit jurnalının başlığı çağıran tərəfdən İXTİYARİ uzunluqda gələ
        # bilər (məs. filtr aralığı adına görə), ona görə burada KƏSİLİR.
        sheet.title = sheet_title[:31]

        header_keys = [key for key, _ in headers]
        self._write_headers(sheet, tuple(label for _, label in headers))
        for row in rows:
            sheet.append([row.get(key, "") for key in header_keys])

        footer_rows = 0
        if note:
            self._add_period_note(sheet, column_count=len(headers), note=note)
            footer_rows = 2
        self._finalize(sheet, len(headers), footer_rows=footer_rows)
        return self._save(workbook, file_name, report="generic_table")

    # ------------------------------- köməkçilər ------------------------------ #

    @staticmethod
    def _write_headers(sheet: Worksheet, headers: tuple[str, ...]) -> None:
        sheet.append(list(headers))
        for cell in sheet[1]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.freeze_panes = "A2"

    @staticmethod
    def _add_period_note(sheet: Worksheet, *, column_count: int, note: str) -> None:
        """Faylın altına izah sətri — faylı açan mühasib məqsədi görsün.

        Başlıqda deyil, ALTDA: yuxarıda olsaydı `SUM()` və filtr aralıqları
        sürüşərdi və mühasibin adi Excel vərdişləri pozulardı.
        """
        sheet.append([])
        sheet.append([note])
        last_row = sheet.max_row
        sheet.merge_cells(
            start_row=last_row, start_column=1, end_row=last_row, end_column=column_count
        )
        sheet.cell(row=last_row, column=1).font = Font(italic=True)

    @staticmethod
    def _finalize(sheet: Worksheet, column_count: int, *, footer_rows: int = 2) -> None:
        """Sütun enlərini məzmuna görə tənzimləyir və avtofiltr qoyur.

        `footer_rows`: `_add_period_note()`-un ƏLAVƏ etdiyi sətir sayı (boş
        sətir + izah sətri = 2) — DEFOLT DƏYİŞMİR, çünki `write_attendance`/
        `write_bonus_penalty` HƏMİŞƏ izah sətri yazır. `write_table()` isə
        `note=""` olanda heç nə əlavə ETMİR (bax modul başlığı, "ÜMUMİ
        CƏDVƏL") — `0` ötürür ki, avtofiltr axırıncı HƏQİQİ məlumat sətrini
        itirməsin.
        """
        for index in range(1, column_count + 1):
            letter = get_column_letter(index)
            longest = max(
                (len(str(cell.value)) for cell in sheet[letter] if cell.value is not None),
                default=_MIN_COLUMN_WIDTH,
            )
            width = min(max(longest + 2, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
            sheet.column_dimensions[letter].width = width

        # Filtr YALNIZ məlumat aralığına qoyulur — izah sətri daxil edilsəydi,
        # süzgəc onu da gizlədə bilərdi.
        data_rows = max(sheet.max_row - footer_rows, 1)
        sheet.auto_filter.ref = f"A1:{get_column_letter(column_count)}{data_rows}"

    def _save(self, workbook: Workbook, filename: str, *, report: str) -> Path:
        try:
            self._output_dir.mkdir(parents=True, exist_ok=True)
            target = self._output_dir / filename
            workbook.save(target)
        except OSError as exc:
            raise ExcelExportError(
                f"Hesabat faylı yazıla bilmədi: {filename}",
                context={"path": str(self._output_dir), "error": str(exc)},
            ) from exc

        _audit_log.info("REPORT_FILE_WRITTEN", extra={"report": report, "path": str(target)})
        return target


__all__ = [
    "ATTENDANCE_HEADERS",
    "BONUS_PENALTY_HEADERS",
    "NOTE_HEADER",
    "ExcelExportError",
    "ExcelReportWriter",
]
